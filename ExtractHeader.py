import pandas as pd
from openpyxl import load_workbook

def extract_section_dataframe(workbook_name: str, sheet_name: str, header_columns: list[str], section: str) -> pd.DataFrame:
    """
    Extracts a DataFrame from an Excel sheet section identified by a super header.
    
    :param workbook_name: Path to the Excel workbook.
    :param sheet_name: Name of the sheet to read from.
    :param header_columns: List of desired column header names (subset of actual headers).
    :param section: The super header text to locate the section.
    :return: DataFrame containing the data from the section with specified columns.
    """
    wb = load_workbook(filename=workbook_name, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    data_rows = []
    in_section = False
    headers_found = False
    header_indices = []
    current_row = 1  # 1-indexed for openpyxl
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    while current_row <= max_row:
        row_values = [ws.cell(row=current_row, column=col).value for col in range(1, max_col + 1)]
        
#        print(f"Row {current_row}: {row_values}")  # Debug: print current row values
        # Strip None and check if row is effectively empty
        stripped_row = [val for val in row_values if val is not None]
        is_empty = len(stripped_row) == 0
        
        if not in_section:
            # Look for super header in the first cell
            if row_values and row_values[0] == section:
                in_section = True
            current_row += 1
            continue
        
        if in_section and not headers_found:
            if is_empty:
                current_row += 1
                continue
            # Get headers and find indices of desired columns
            actual_headers = [str(val).strip() if val else '' for val in row_values]
            header_indices = []
            for col in header_columns:
                try:
                    idx = actual_headers.index(col)
                    header_indices.append(idx)
                except ValueError:
                    raise ValueError(f"Column '{col}' not found in section '{section}' headers: {actual_headers}")
            headers_found = True
            current_row += 1
            continue
        
        if headers_found:
            if is_empty or (row_values and row_values[0] and isinstance(row_values[0], str) and any(sh in row_values[0] for sh in ["Equity", "Mutual Funds"])):
                break  # End of section
            # Collect data for specified columns
            data_row = [row_values[idx] if idx < len(row_values) else None for idx in header_indices]
            data_rows.append(data_row)
            current_row += 1
            continue
        
        current_row += 1
    
    if not headers_found:
        raise ValueError(f"Section '{section}' not found or headers missing.")
    
    df = pd.DataFrame(data_rows, columns=header_columns)
    return df